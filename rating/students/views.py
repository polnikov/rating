import concurrent.futures
import logging

from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.shortcuts import redirect, render, get_object_or_404
from django.urls import reverse
from django.utils.translation import gettext as _
from django.views.generic import (
    CreateView, DeleteView, DetailView, ListView, UpdateView, View, TemplateView,
)

from groups.models import Group
from students.forms import ResultForm, StudentForm
from students.models import Result, Semester, Student, StudentLog
from subjects.models import Cathedra, GroupSubject, Subject
from rating.functions import calculate_rating


logger = logging.getLogger(__name__)


class StudentView(LoginRequiredMixin, TemplateView):
    template_name = 'students/students.html'


    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['students'] = Student.active_objects.all().count()
        context['form'] = StudentForm()
        return context


class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student

    def get(self, request, pk, **kwargs):
        student = get_object_or_404(Student.objects.select_related('group', 'semester', 'basis'), student_id__exact=pk)

        try:
            history = StudentLog.objects.select_related('user').filter(
                record_id=student.student_id).order_by('-timestamp').values()
        except Exception as ex:
            logger.info(f'[-] История изменений по студенту {student} отсутствует', extra={'Exception': ex})
            history = 'Error'

        # all student marks
        marks = Result.objects.select_related().filter(
            students=student.student_id).filter(
            ~Q(groupsubject__subjects__form_control__exact='Зачет'))
        # all group results exclude non-nu,eric marks
        atts = GroupSubject.objects.select_related('subjects'
            ).filter(groups=student.group
            ).filter(~Q(subjects__form_control__exact='Зачет'))

        # calculation of the average ranking for semesters and the total
        rating_by_semester_bac = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
            5: 0,
            6: 0,
            7: 0,
            8: 0,
        }
        rating_by_semester_mag = {
            1: 0,
            2: 0,
            3: 0,
            4: 0,
        }
        all_num_marks = []
        all_marks = []
        if student.level == Student.Level.BAC:
            semesters = range(1, 9)
            rating_by_semester = rating_by_semester_bac
        elif student.level == Student.Level.MAG:
            semesters = range(1, 5)
            rating_by_semester = rating_by_semester_mag

        for i in semesters:
            # all marks per semester
            sem_marks_all = marks.select_related('subjects').filter(
                groupsubject__subjects__semester__semester=i
            ).values('mark')
            # take only the last mark and exclude <ня> и <2>
            sem_marks = list(filter(lambda x: x not in ['ня', '2'], [i['mark'][-1] for i in sem_marks_all]))
            all_marks += sem_marks
            # quantity of groupsubjects with marks per semester
            num_atts = atts.filter(subjects__semester=i).count()
            all_num_marks.append(num_atts)
            # quantity of each mark <3 | 4 | 5>
            count_marks = dict(Counter(sem_marks))

            # calculate average ranking per semester
            if num_atts:
                sem_rating = round(sum([int(k)*v for k, v in count_marks.items()]) / num_atts, 2)
            else:
                sem_rating = 0

            rating_by_semester[i] = sem_rating

        # calculate total ranking
        if sum(all_num_marks):
            rating = round(sum(list(map(int, (all_marks)))) / sum(all_num_marks), 2)
        else:
            rating = 0

        form = StudentForm()

        context = {
            'student': student,
            'history': history,
            'rating': rating,
            'rating_by_semester': rating_by_semester,
            'form': form,
            'result_form': ResultForm(),
        }
        return render(request, 'students/student_detail.html', context=context)


class StudentRatingTableView(LoginRequiredMixin, ListView):
    """Head of table of the students average ranking. Semesters"""
    def get(self, request):
        semesters = Semester.objects.all()
        groups = Group.objects.filter(is_archived=False)
        return render(request, 'students/students_rating.html', context={'semesters': semesters,
                                                                         'groups': groups})


class StudentRatingApiView(LoginRequiredMixin, View):
    """Calculating of the students average ranking."""
    def get(self, request):
        logger.info(f'Расчет среднего балла...')
        serialized_data = []
        sem_start = request.GET.get('semStart', '')
        sem_stop = request.GET.get('semStop', '')
        groups = request.GET.getlist('groups[]', False)

        if sem_start:
            start = sem_start
        else:
            start = 1

        if groups:
            students = Student.active_objects.select_related('group', 'semester', 'basis').filter(
                group__name__in=groups, semester__semester__gte=start)
        else:
            students = Student.active_objects.select_related(
                'group', 'semester', 'basis').filter(
                semester__semester__gte=start)

        flag_1 = not sem_start and not sem_stop
        flag_2 = sem_start and not sem_stop
        flag_3 = not sem_start and sem_stop
        flag_4 = sem_start and sem_stop == '-'

        if flag_1 or flag_2 or flag_3 or flag_4:
            logger.info(f'|---> Расчет среднего балла за семестр {start} для студентов группы {groups}')

            with concurrent.futures.ThreadPoolExecutor() as executor:
                # the students average ranking for specified semester. default - for 1t
                future_to_student = {
                    executor.submit(calculate_rating, student, start): student
                    for student in students
                }
                for future in concurrent.futures.as_completed(future_to_student):
                    student = future_to_student[future]
                    try:
                        rating = future.result()
                    except Exception as exc:
                        logger.error(f'calculation of student {student.student_id} rating failed: {exc}')
                    else:
                        serialized_data.append({
                            'studentId': student.student_id,
                            'fullname': student.fullname,
                            'group': student.group.name,
                            'currentSemester': student.semester.semester,
                            'basis': student.basis.name,
                            'level': student.level,
                            'rating': rating,
                            'isIll': student.is_ill,
                            'tag': student.tag,
                        })
        else:
            # the students average ranking for specified period
            start, stop = sem_start, sem_stop
            logger.info(f'Расчет среднего балла за период с {start} по {stop} семестр для студентов группы {groups}')

            with concurrent.futures.ThreadPoolExecutor() as executor:
                future_to_student = {
                    executor.submit(calculate_rating, student, start, stop): student
                    for student in students
                }
                for future in concurrent.futures.as_completed(future_to_student):
                    student = future_to_student[future]
                    try:
                        rating = future.result()
                    except Exception as exc:
                        logger.error(f'calculation of student {student.student_id} rating failed: {exc}')
                    else:
                        serialized_data.append({
                            'studentId': student.student_id,
                            'fullname': student.fullname,
                            'group': student.group.name,
                            'currentSemester': student.semester.semester,
                            'basis': student.basis.name,
                            'level': student.level,
                            'rating': rating,
                            'isIll': student.is_ill,
                            'tag': student.tag,
                        })
        serialized_data = sorted(serialized_data, key=lambda d: d['rating'])

        return JsonResponse({'data': serialized_data})


class GraduatesView(LoginRequiredMixin, TemplateView):
    template_name = 'students/graduates.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        graduates = Student.archived_objects.filter(status='Выпускник')
        context['graduates'] = graduates
        return context


class StudentsMoneyListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'students/students_money.html'
    queryset = Student.active_objects.select_related('basis', 'group', 'semester')


class StudentsDebtsListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'students/students_debts.html'

    def get_queryset(self):
        negative = ['ня', 'нз', '2']

        # students ids who has debts
        negative_students = Result.objects.select_related('students').filter(
            mark__contained_by=negative, is_archived=False).values('students__student_id')
        # filter students by ids
        students = Student.active_objects.select_related(
            'basis', 'group', 'semester').filter(
            student_id__in=negative_students)

        for st in students:
            all_marks = [
                i[0]
                for i in st.result_set.select_related().filter(
                    groupsubject__subjects__semester__semester=st.semester.semester, groupsubject__groups__name=st.group.name,
                    mark__contained_by=negative).values_list('mark')]
            marks_att1 = [i[0] for i in all_marks]
            marks_att2 = [i[1] for i in list(filter(lambda x: len(x) in [2, 3], all_marks))]
            marks_att3 = [i[2] for i in list(filter(lambda x: len(x) == 3, all_marks))]
            count_marks_att1 = dict(Counter(marks_att1))
            count_marks_att2 = dict(Counter(marks_att2))
            count_marks_att3 = dict(Counter(marks_att3))
            st.att1 = sum(list(map(lambda x: count_marks_att1.get(x, 0), negative)))
            st.att2 = sum(list(map(lambda x: count_marks_att2.get(x, 0), negative)))
            st.att3 = sum(list(map(lambda x: count_marks_att3.get(x, 0), negative)))

        return students


def search_results(request):
    if request.method == 'GET':
        search = request.GET.get('search')
        students = Student.objects.filter(
            Q(student_id__icontains=search) |
            Q(last_name__icontains=search) |
            Q(first_name__icontains=search) |
            Q(second_name__icontains=search) |
            Q(comment__icontains=search) |
            Q(tag__icontains=search)
        )
        subjects = Subject.objects.filter(name__icontains=search)
        groupsubjects = GroupSubject.objects.filter(
            Q(subjects__name__icontains=search) |
            Q(teacher__icontains=search) |
            Q(comment__icontains=search)
        )
        context = {
            'search': search,
            'students': students,
            'subjects': subjects,
            'groupsubjects': groupsubjects,
        }
    return render(request, 'search_results.html', context=context)


def download_excel_data(request):
    negative = ['ня', 'нз', '2']
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="Zadolzhennosti FIEGH.xlsx"'

    att_level = request.GET.get('att')
    logger.info(f'Формирование excel файла с задолженностями за {att_level} период аттестации...')

    # groupsubjects ids with negative results per att period
    match att_level:
        case 'att1':
            negative_groupsubjects_ids = Result.objects.select_related().filter(
                Q(mark__len=1) & Q(mark__0__in=negative)).values('groupsubject__id')
        case 'att2':
            negative_groupsubjects_ids = Result.objects.select_related().filter(
                Q(mark__len=2) & Q(mark__1__in=negative)).values('groupsubject__id')
        case 'att3':
            negative_groupsubjects_ids = Result.objects.select_related().filter(
                mark__2__in=negative).values('groupsubject__id')

    if not negative_groupsubjects_ids:
        url = reverse('students:debts')
        logger.info(f'|---> Задолженностей за {att_level} период аттестации нет')
        return HttpResponseRedirect(url)
    else:
        # groupsubjects
        group_subjects = GroupSubject.active_objects.select_related().filter(id__in=negative_groupsubjects_ids)
        
        # cathedras
        cathedras = list(set(group_subjects.values_list('subjects__cathedra__short_name', flat=True)))

        # create xls file
        book = Workbook()
        # headers names
        header = ['Дисциплина', 'Форма контроля', 'Группа', 'Семестр', 'Преподаватель', 'Студенты']

        for cathedra in cathedras:
            # create the sheet with the name equal cathedra short name
            book.create_sheet(title=cathedra, index=None)
            sheet = book[cathedra]
            sheet.column_dimensions['A'].width = 75
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 15
            sheet.column_dimensions['E'].width = 30
            sheet.column_dimensions['F'].width = 150
            sheet.auto_filter.ref = 'A1:F1'

            # write the header line on the sheet
            for col in range(1, len(header) + 1):
                cell = sheet.cell(row=1, column=col, value=header[col - 1])
                sheet[cell.coordinate].font = Font(bold=True, size=14)

            # groupsubjects for the cathedra
            cathedra_group_subjects = group_subjects.filter(subjects__cathedra__short_name=cathedra)

            row = 2
            for gs in cathedra_group_subjects:
                # negative results per cathedra groupsubject
                match att_level:
                    case 'att1':
                        res = gs.result_set.select_related().filter(mark__0__in=negative)
                    case 'att2':
                        res = gs.result_set.select_related().filter(mark__1__in=negative)
                    case 'att3':
                        res = gs.result_set.select_related().filter(mark__2__in=negative)

                # students full name with debts
                students = res.values_list('students__last_name', 'students__first_name', 'students__second_name')
                students = ', '.join([' '.join(i) for i in students])
                res_data = [
                    gs.subjects.name,
                    gs.subjects.form_control,
                    gs.groups.name,
                    gs.subjects.semester.semester,
                    gs.teacher,
                ]
                res_data.append(students)

                for col, data in enumerate(res_data):
                    sheet.cell(row=row, column=col + 1, value=data)
                    sheet.cell(row=row, column=col + 1).font = Font(size=12)
                row += 1

        # create sheet with contents
        sheet = book['Sheet']
        sheet.title = 'Оглавление'
        sheet.column_dimensions['A'].width = 200

        sheets = book.sheetnames
        sheets = sheets[1:]
        row = 1
        for sh in sheets:
            link = f'#{sh}!A1'
            sheet.cell(row=row, column=1, value=Cathedra.objects.get(short_name=sh).name).hyperlink = link
            sheet.cell(row=row, column=1).font = Font(size=14)
            row += 1

        book.save(response)
        logger.info(f'|---> [+] Файл с задолженностями за {att_level} период аттестации успешно сформирован')
        return response
